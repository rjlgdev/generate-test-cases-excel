import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_feature_file(file_path):
    """
    Analisa o arquivo .feature para extrair informações de Casos de Teste.
    """
    test_cases = []
    current_scenario = {}
    
    # Regex para capturar os dados
    feature_re = re.compile(r"^\s*Feature:\s*(.*)")
    scenario_re = re.compile(r"^\s*Scenario:\s*(.*)")
    step_re = re.compile(r"^\s*(Given|When|Then|And|But)\s+(.*)")
    evidence_re = re.compile(r"^\s*#Evidência:\s*(.*)")
    result_re = re.compile(r"^\s*#Resultado:\s*(.*)")
    
    # Variáveis para armazenar dados da Feature e do Scenario atual
    feature_name = ""
    current_evidence = ""
    current_result = ""
    
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        
    for line in lines:
        line = line.strip()
        
        # 1. Capturar o nome da Feature
        match_feature = feature_re.match(line)
        if match_feature:
            feature_name = match_feature.group(1).strip()
            continue
            
        # 2. Capturar Evidência e Resultado (estão como comentários antes do Scenario)
        match_evidence = evidence_re.match(line)
        if match_evidence:
            current_evidence = match_evidence.group(1).strip()
            continue
            
        match_result = result_re.match(line)
        if match_result:
            current_result = match_result.group(1).strip()
            continue
            
        # 3. Capturar o início do Scenario
        match_scenario = scenario_re.match(line)
        if match_scenario:
            # Se já houver um cenário anterior, salve-o
            if current_scenario:
                test_cases.append(current_scenario)
            
            # Iniciar novo cenário
            scenario_title = match_scenario.group(1).strip()
            
            # Tentar extrair o ID do Caso de Teste (ex: "01) Cadastrar Primeira Fazenda" -> "TC_01")
            tc_id_match = re.search(r"(\d+)\)", scenario_title)
            tc_id = f"TC_{tc_id_match.group(1).zfill(2)}" if tc_id_match else f"TC_{len(test_cases) + 1:02d}"
            
            # O Test Case (coluna) será o título do Scenario
            test_case_col = scenario_title
            
            current_scenario = {
                "id": tc_id,
                "scenario": feature_name, # Usar o nome da Feature como Test Scenario
                "test_case": test_case_col,
                "pre_condition": [],
                "steps": [], # Armazenará tuplas (keyword, text)
                "expected_result": [], # Armazenará tuplas (keyword, text)
                "evidence": current_evidence,
                "status": current_result,
                "has_when": False # Flag para saber se já encontramos um 'When'
            }
            
            # Resetar Evidência e Resultado para o próximo cenário
            current_evidence = ""
            current_result = ""
            continue
            
        # 4. Capturar os Steps (Given, When, Then, And, But)
        match_step = step_re.match(line)
        if match_step and current_scenario:
            step_type = match_step.group(1)
            step_text = match_step.group(2).strip()
            
            # Se for 'When', marca que os próximos 'Given'/'And' não são mais pré-condições
            if step_type == "When":
                current_scenario["has_when"] = True
                
            # Lógica para separar Pre-Condition, Test Steps e Expected Result
            if step_type in ["Given", "And"] and not current_scenario["has_when"]:
                # Given/And antes de When são Pre-Conditions
                current_scenario["pre_condition"].append((step_type, step_text))
            elif step_type in ["When", "And", "But"] or (step_type == "Given" and current_scenario["has_when"]):
                # When/And/But são Test Steps. Given após When também é Test Step.
                current_scenario["steps"].append((step_type, step_text))
            elif step_type == "Then":
                # Then é o Expected Result
                current_scenario["expected_result"].append((step_type, step_text))
                
    # Adicionar o último cenário
    if current_scenario:
        test_cases.append(current_scenario)
        
    return test_cases, feature_name

def create_excel_sheet(test_cases, feature_name, output_file="Casos_de_Teste.xlsx"):
    """
    Cria e formata a planilha Excel com os dados extraídos.
    """
    # 1. Configuração do Workbook e Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = feature_name if feature_name else "Casos de Teste"
    
    # 2. Estilos
    # Fonte e Preenchimento do Cabeçalho (Azul Escuro)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    
    # Fonte e Preenchimento do Título (Azul Claro)
    title_font = Font(bold=True, color="000000")
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Estilo de Borda
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Alinhamento
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 3. Cabeçalho de Informações do Projeto (Baseado na Imagem)
    info_headers = ["Project Name:", "Module Name:", "Reference Document:", "Created by:"]
    info_values = [feature_name, feature_name, "Feature File", "Rafael La Guardia"]
    
    for i, (header, value) in enumerate(zip(info_headers, info_values)):
        row = i + 1
        ws[f'A{row}'] = header
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].border = thin_border
        
        # Mesclar células para as informações do projeto
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        
    # 4. Cabeçalho da Tabela de Casos de Teste
    header_row = len(info_headers) + 2 # Começa após as informações do projeto
    
    columns = [
        "Test Case ID", "Test Scenario", "Test Case", "Pre-Condition", 
        "Test Steps", "Expected Result", "Evidência", "Status"
    ]
    
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}{header_row}']
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # 5. Preencher os Dados dos Casos de Teste
    row_num = header_row + 1
    
    for tc in test_cases:
        # Formatar listas em strings com quebras de linha e numeração
        # Pré-condição: Apenas o texto, sem numeração, mas com a palavra-chave Gherkin
        pre_condition_str = "\n".join([f"{pc[0]} {pc[1]}" for pc in tc["pre_condition"]])
        
        # Test Steps: Com numeração e a palavra-chave Gherkin
        steps_str = "\n".join([f"{i+1}. {step[0]} {step[1]}" for i, step in enumerate(tc["steps"])])
        
        # Expected Result: Com numeração e a palavra-chave Gherkin
        expected_result_str = "\n".join([f"{i+1}. {er[0]} {er[1]}" for i, er in enumerate(tc["expected_result"])])
        
        data = [
            tc["id"],
            tc["scenario"],
            tc["test_case"],
            pre_condition_str,
            steps_str,
            expected_result_str,
            tc["evidence"],
            tc["status"]
        ]
        
        for col_num, value in enumerate(data, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}{row_num}']
            cell.value = value
            cell.border = thin_border
            cell.alignment = top_left_align
            
            # Ajuste de cor para a coluna Status
            if col_num == 8: # Coluna Status
                if "SUCESSO" in str(value).upper():
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde Claro
                elif "FALHA" in str(value).upper():
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Vermelho Claro
                
        row_num += 1
        
    # 6. Ajustar Largura das Colunas
    column_widths = [15, 25, 35, 30, 40, 40, 30, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        
    # 7. Salvar o arquivo
    wb.save(output_file)
    return output_file

if __name__ == "__main__":
    # O script espera o caminho do arquivo .feature como argumento
    import sys
    if len(sys.argv) < 2:
        print("Uso: python generate-test-case.py <caminho_para_o_arquivo.feature>")
        sys.exit(1)
        
    feature_file_path = sys.argv[1]
    
    try:
        print(f"Analisando arquivo: {feature_file_path}...")
        test_cases, feature_name = parse_feature_file(feature_file_path)
        
        if not test_cases:
            print("Nenhum cenário de teste encontrado.")
            sys.exit(0)
            
        output_filename = f"{feature_name.replace(' ', '_')}_Casos_de_Teste_v2.xlsx" if feature_name else "Casos_de_Teste_v2.xlsx"
        
        print(f"Gerando planilha Excel: {output_filename}...")
        output_path = create_excel_sheet(test_cases, feature_name, output_filename)
        print(f"Planilha gerada com sucesso em: {output_path}")
        
    except FileNotFoundError:
        print(f"Erro: Arquivo não encontrado em {feature_file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        sys.exit(1)
